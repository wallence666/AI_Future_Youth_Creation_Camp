# -*- coding: utf-8 -*-
"""
build_model_data.py
從澳門旅遊局分區訪客統計 (data/dst_visitor_01_stats.xlsx) 生成前端模型數據：
  app/data/model.json
    - districts: 各區基準熱度 B(s)（歸一化 [0,1]）與 24 小時平均人流曲線
    - curves:    依地點類型分群的時段曲線 T_c(s,t)（4 群，縮放至 [0,1.2]）
                 分群依據 docs/01 技術文檔 §3.1：廟宇廣場型／商業步行街型／室內場館型／濱海戶外型
用法：python tools/build_model_data.py
"""
import json
import math
import os
import shutil
from collections import defaultdict

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "data", "dst_visitor_01_stats.xlsx")
OUT = os.path.join(ROOT, "app", "data", "model.json")

# 25 區 -> 5 類時段曲線分群（依據各區主要場景屬性）
# v1.7：路氹填海區（澳門威尼斯人/銀河/倫敦人所在地）獨立為 resort_247，不再與其他室內場館
# 混合計算——三大博企度假村是真正的 24 小時全天候營運場館，其真實 dst_visitor 數據（谷值/峰值
# 比例約 0.68）遠高於「科學館」「旅遊塔」這類有固定開放時間的室內場館，混在同一群組計算會讓
# 曲線形狀被互相稀釋失真；獨立分群後兩邊都能各自忠實反映真實數據，且色帶正規化基準（見
# model.js computeGroupBounds()）改為按群組獨立計算，從根源解決度假村類景點「長期偏紅、看不出
# 時段差異」的問題，見 docs/01 版本修訂摘要 v1.7。
DISTRICT_GROUP = {
    "沙梨頭及大三巴區": "temple_plaza",   # 大三巴、廟宇廣場
    "中區": "temple_plaza",               # 議事亭前地廣場
    "下環區": "temple_plaza",             # 媽閣廟一帶
    "氹仔舊城及馬場區": "pedestrian",     # 官也街步行街
    "荷蘭園區": "pedestrian",             # 塔石/荷蘭園商業街
    "新橋區": "pedestrian",               # 三盞燈民生商業
    "高士德及雅廉訪區": "pedestrian",
    "路氹填海區": "resort_247",           # 金光大道三大博企度假村（24 小時全天候營運，獨立分群）
    "氹仔中心區": "indoor",               # 其餘室內場館型（有固定開放時間）
    "外港及南灣湖新填海區": "indoor",     # 外港碼頭/漁人碼頭綜合體
    "新口岸區": "indoor",                 # 皇朝/新口岸酒店娛樂區
    "路環區": "waterfront",               # 黑沙海灘、路環濱海
    "南西灣及主教山區": "waterfront",     # 南灣湖、主教山觀景
    "東望洋區(松山區)": "waterfront",     # 東望洋燈塔戶外
    "海洋及小潭山區": "waterfront",       # 小潭山觀景台
}
GROUP_LABEL = {
    "temple_plaza": "廟宇廣場型",
    "pedestrian": "商業步行街型",
    "indoor": "室內場館型",
    "waterfront": "濱海戶外型",
    "resort_247": "度假村全天型",
}
CURVE_MAX = 1.2  # 文檔規定 T_c 值域 [0, 1.2]


def load_profiles():
    wb = openpyxl.load_workbook(XLSX)
    ws = wb[wb.sheetnames[0]]
    per_district_hour = defaultdict(lambda: defaultdict(list))
    for row in ws.iter_rows(min_row=2, values_only=True):
        _year, _month, period, dcn, _dpt, _den, hc, _notes = row
        if period is None or hc is None or dcn is None:
            continue
        hour = int(str(period).split(":")[0])
        per_district_hour[dcn][hour].append(hc)
    # 每區 24 小時平均人流
    profiles = {}
    for d, hours in per_district_hour.items():
        profiles[d] = [sum(hours[h]) / len(hours[h]) if hours[h] else 0.0 for h in range(24)]
    return profiles


def main():
    profiles = load_profiles()

    # B(s)：各區日平均人流，以最大區歸一化。
    # 熱力圖表達「擁擠體感」（密度）而非絕對客流，否則路氹度假區的龐大吞吐量
    # 會令大三巴等步行區永遠排不到前面；故以開方壓縮量綱差異（同時保留 raw 值備查）。
    daily_avg = {d: sum(v) / 24.0 for d, v in profiles.items()}
    max_avg = max(daily_avg.values())
    districts = {
        d: {
            "base": round(math.sqrt(daily_avg[d] / max_avg), 4),
            "baseRaw": round(daily_avg[d] / max_avg, 4),
            "group": DISTRICT_GROUP.get(d, "pedestrian"),
            "hourly": [round(x, 1) for x in profiles[d]],
        }
        for d in profiles
    }

    # T_c(s,t)：同群各區曲線先各自歸一化再取均值，最後縮放到 [0, 1.2]（峰值對齊 CURVE_MAX）
    grouped = defaultdict(list)
    for d, prof in profiles.items():
        g = DISTRICT_GROUP.get(d)
        if not g:
            continue
        peak = max(prof) or 1.0
        grouped[g].append([x / peak for x in prof])
    curves = {}
    for g, mats in grouped.items():
        avg = [sum(m[h] for m in mats) / len(mats) for h in range(24)]
        peak = max(avg) or 1.0
        curves[g] = {
            "label": GROUP_LABEL[g],
            "values": [round(x / peak * CURVE_MAX, 4) for x in avg],
        }

    # v1.7：不再對曲線形狀做人工振幅校正（v1.6 一度用「拉齊谷值/峰值比例」的方式修正 indoor
    # 群組，事後證實治標不治本——即使曲線形狀壓平，度假村類景點的 B(s) 基準熱度仍遠高於其他
    # 景點，套用「全部 18 個景點共用一組正規化色帶基準」時依然會被判定長期偏紅）。
    # 真正的根因是「不同類型景點的絕對人流量綱本來就不同，硬要共用同一把色帶尺規」，故改為：
    #   1) 路氹填海區（三大博企度假村）獨立分群為 resort_247，忠實呈現其真實 24 小時曲線
    #      （不再與科學館、旅遊塔等有固定開放時間的室內場館混合計算，見上方 DISTRICT_GROUP 註解）；
    #   2) app/js/model/model.js 的色帶正規化基準改為「按 T_c 分群各自獨立計算」
    #      （computeGroupBounds()，非全體 18 景點共用一組 min/max），
    #      每個分群色帶都反映「該類型景點自己一天內的相對忙閒」，不再被其他類型的量綱壓縮。
    # 曲線本身維持忠於真實 dst_visitor 數據，不做人工形狀調整，見 docs/01 版本修訂摘要 v1.7。

    out = {
        "meta": {
            "source": "澳門旅遊局分區訪客統計 dst_visitor_01_stats.xlsx（2026-06，逐時）",
            "weights": {"w1": 0.6, "w2": 0.4},
            "curveMax": CURVE_MAX,
            "generated_by": "tools/build_model_data.py",
        },
        "districts": districts,
        "curves": curves,
    }
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=1)

    # 同步爬蟲活動數據 -> app 自帶一份，作為 H(t) 放大係數輸入
    events_src = os.path.join(ROOT, "data", "events.json")
    if os.path.exists(events_src):
        shutil.copyfile(events_src, os.path.join(os.path.dirname(OUT), "events.json"))
        print("copied: data/events.json -> app/data/events.json")
    print("written:", OUT)
    print("districts:", len(districts), "| curves:", list(curves))
    for g, c in curves.items():
        print(g, c["label"], "peak_hour=", c["values"].index(max(c["values"])), "max=", max(c["values"]))


if __name__ == "__main__":
    main()
